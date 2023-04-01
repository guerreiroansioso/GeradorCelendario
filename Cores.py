from openpyxl.styles import PatternFill

class Seletor:
    @staticmethod
    def corDias(dia):
        # Cores básicas de fundo.
        corAzul     = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        corLaranja  = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        corVerde    = PatternFill(start_color='C5E0B3', end_color='C5E0B3', fill_type='solid')
        corRoxo     = PatternFill(start_color='CBA6f0', end_color='CBA6f0', fill_type='solid')
        corAmarelo  = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

        # Cor padrão.
        corBranco   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Implementação de switch para selecionar cores adequadas.
        switcher = {
            0: corVerde,
            1: corLaranja,
            2: corAzul,
            3: corVerde,
            4: corLaranja,
            5: corAzul,
            6: corRoxo,
            "padrão": corBranco
        }
        return switcher.get(dia, switcher['padrão'])