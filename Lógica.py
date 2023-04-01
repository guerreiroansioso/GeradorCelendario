import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

# Enumeração para dias da semana.
diasPtBr = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
diasEnUs = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

corAzul     = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
corLaranja  = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
corVerde    = PatternFill(start_color='C5E0B3', end_color='C5E0B3', fill_type='solid')
corRoxo     = PatternFill(start_color='CBA6f0', end_color='CBA6f0', fill_type='solid')
corAmarelo  = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

corBranco   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

def corSeletorDias(diaNumero):
    switcher = {
        0: corVerde,
        1: corLaranja,
        2: corAzul,
        3: corVerde,
        4: corLaranja,
        5: corAzul,
        6: corRoxo,
        "default": corBranco
    }
    return switcher.get(diaNumero, switcher['default'])

# Criar solução excel.
arquivoExcel = openpyxl.Workbook()
arquivoExcel.title = 'Calendário da fusão'

# Criar planilha dentro da solução.
fusãoAtual = arquivoExcel.active
fusãoAtual.title = 'Fusão'

# Fixar a primeira linha.
fusãoAtual.freeze_panes = 'A2'

# Retornar os dias da semana.
dataInicial = '2023-04-02'
dataFinal = '2023-04-05'

diaInicio = datetime.strptime(dataInicial, '%Y-%m-%d')
diaFinal = datetime.strptime(dataFinal, '%Y-%m-%d')
diasQtd = (diaFinal - diaInicio).days

listaData = []
listaDia = []
for acréscimo in range(diasQtd):
    data = diaInicio + timedelta(days = acréscimo)
    dia = diasPtBr[data.weekday()]
    listaData.append(data)
    listaDia.append(dia)

for valor in range(diasQtd):
    fusãoAtual.cell(row = 1, column = valor + 1).value = listaDia[valor]
    fusãoAtual.cell(row = 2, column = valor + 1).value = listaData[valor]
    fusãoAtual.cell(row = 1, column = valor + 1).fill = corSeletorDias(listaData[valor].weekday())

# Write some data into cells
fusãoAtual['A5'] = 'Exemplo'

# Savar solução.
print('Salvando arquivo...')
arquivoExcel.save('Calendário.xlsx')
print('Calendário salvo em "Calendário.xlsx".')